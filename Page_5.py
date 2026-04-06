import streamlit as st
from datetime import datetime, timedelta, date
from Category import return_all_categories
from Day import return_all_days, create_day
from General_Functions import return_table
from Ingredient import return_all_ingredients
from Meal import validate_meal, meal_name_label, return_all_meals, meal_id_to_index
from MealCombination import return_all_meal_combinations
from MealType import return_all_meal_types
from Menu import menu
from Page_14 import alter_schedule_officially, remove_schedule_officially, add_schedule_officially
from Page_6 import get_table
from Rule import return_all_rules
from Schedule import return_all_schedules, outcome_table_with_tags, update_schedule
from User import validate_user
import io
from openpyxl import Workbook
from collections import defaultdict
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Step 1: Initialize session state variables (first run only)

if 'page' not in st.session_state:
    st.session_state.page = 1  # Default page to load

if 'error_status' not in st.session_state:
    st.session_state.error_status = None  # Tracks whether an error message should be shown

if 'current_user' not in st.session_state:
    st.session_state.current_user = None  # Stores the CodeID of the signed-in user

if 'error' not in st.session_state:
    st.session_state.error = 'You are doing great! Keep going.'  # Stores the current error message


def page_5_layout():
    # Step 1: Validate current user
    # Ensures the user exists and is allowed to access this page.
    message, entry, status = validate_user(st.session_state.current_user)

    if status:
        # Step 2: Render the page layout
        # Show navigation and page heading for the authenticated user.

        # Step 2a: Render main navigation menu
        menu(entry)

        # Step 2b: Render page title
        st.title("My Meal Plan")

        # Step 3: Collect the dates the user wants to plan meals for
        st.header("Date Selection")
        date_table = get_table()

        if len(date_table) >= 1:

            # Step 4: Ensure every selected date has a Day entry in storage
            # Convert the selected dates into a list of dictionaries with Date and CodeID.
            match_date_with_ID = generate_missing_days(date_table)

            if len(match_date_with_ID) >= 1:

                # Step 5: Render meal schedule display options
                # User can choose whether to view empty slots, existing slots, or both.
                st.header("Meal Schedule")
                add_slot_column, alter_slot_column = st.columns(2, vertical_alignment="center")

                with add_slot_column:
                    add_slot = st.checkbox("Show Empty Slots")

                with alter_slot_column:
                    alter_slot = st.checkbox("Show Existing Slots")

                # Step 6: Retrieve and sort the user's meal types
                # Meal types are displayed in ascending priority order.
                meal_types = return_all_meal_types({'UserID': st.session_state.current_user})
                meal_types = sorted(meal_types, key=lambda x: x['Priority'])

                if len(meal_types) == 0:
                    # Step 6a: Show message if the user has no meal types
                    # Without meal types, no schedule slots can be displayed or created.
                    st.write(
                        "You seem to have no categories. Structure, even theoretical, saves time. "
                        "Please go to the home page and select one of our pre-made meals to create "
                        "a category or navigate to the profile page to create your own."
                    )

                # Step 6b: Show Rule Violations if they exist
                table, excel_data, filename = find_errors(date_table)
                if len(table) >= 1:
                    with st.container(border=True):
                        st.write(f'We see {len(table)} rule violations:')
                        for entry in table:
                            st.write(f'[!] {entry}')

                # Step 7: Render the selected days in rows of up to four columns
                # Each day column contains all relevant meal slots for that date.
                table = []
                counter = 0
                pointer = 0

                for entry in match_date_with_ID:
                    counter += 1
                    table.append(entry)

                    if counter == 3:
                        cols = st.columns(len(table))
                        for i, val in enumerate(table):
                            with cols[i]:
                                pointer = show_day(table[i], meal_types, pointer, add_slot, alter_slot)
                        table = []
                        counter = 0

                # Step 7a: Render any remaining days that did not fill a row with four
                if len(table) >= 1:
                    cols = st.columns(len(table))
                    for i, val in enumerate(table):
                        with cols[i]:
                            pointer = show_day(table[i], meal_types, pointer, add_slot, alter_slot)

                # Step 9: Download Meal Plan
                st.download_button(
                    label="Download schedule file",
                    data=excel_data,
                    use_container_width=True,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # Step 9: Update schedule statuses based on the current date
        # If meals have passed, direct the user toward the accountability page.
        counter = update_statuses()
        if counter >= 1:
            st.header('Go to Accountability Page')
            st.write(
                f'Congratulations! You have passed {counter} meals of the ones here. '
                f'Go to your accountability page and report if you reached your goals!'
            )
            st.write("It is ok if you didn't, Life happens sometimes")

    else:
        # Step 9: Handle validation failure
        # Store the error so the global error display system can show it elsewhere.
        st.session_state.error, st.session_state.error_status = message, status


def generate_missing_days(date_table):
    # Step 1: Prepare output list
    # This will hold the selected dates matched with their Day CodeIDs.
    match_date_with_ID = []

    # Step 2: Process each selected date
    for entry in date_table:
        # Step 2a: Check whether the date already exists in storage
        date_entry = return_all_days({'Date': entry})

        if len(date_entry) == 1:
            # Step 2b: Reuse the existing Day entry
            match_date_with_ID.append({
                'Date': entry,
                'CodeID': date_entry[0]['CodeID']
            })

        elif len(date_entry) == 0:
            # Step 2c: Create a new Day entry if one does not already exist
            message, date_entry, status = create_day(entry, st.session_state.current_user)

            if not status:
                # Step 2d: Store creation error for global error display
                st.session_state.error, st.session_state.error_status = message, status
            else:
                # Step 2e: Add the newly created Day entry to the output list
                match_date_with_ID.append({
                    'Date': entry,
                    'CodeID': date_entry['CodeID']
                })

    # Step 3: Return all matched dates with IDs
    return match_date_with_ID


def find_errors(date_table):
    # Step 1: Gather all linked data needed to check rule violations for the selected dates.
    categories, days, ingredient_ids_by_meal_id, ingredients, meal_types, meals, rules, schedules = gather_data(
        date_table
    )

    # Step 2: Build one appearance table that tracks every rule-bearing CodeID and the dates it appears on.
    code_appearances = {}

    date_by_day_id = {day["CodeID"]: day["Date"] for day in days}

    code_appearances = get_meal_type_dates(meal_types, schedules, code_appearances, date_by_day_id)
    code_appearances = get_meal_dates(meals, schedules, code_appearances, date_by_day_id)
    code_appearances = get_category_dates(categories, meals, schedules, code_appearances, date_by_day_id)
    code_appearances = get_ingredient_dates(
        ingredients,
        ingredient_ids_by_meal_id,
        schedules,
        code_appearances,
        date_by_day_id
    )

    # Step 3: Build lookup tables for error processing and Excel export.
    item_table = {
        "MealType": {entry["CodeID"]: entry for entry in meal_types},
        "Meal": {entry["CodeID"]: entry for entry in meals},
        "Category": {entry["CodeID"]: entry for entry in categories},
        "Ingredient": {entry["CodeID"]: entry for entry in ingredients}
    }

    day_lookup = {entry["CodeID"]: entry for entry in days}
    meal_type_lookup = {entry["CodeID"]: entry for entry in meal_types}
    meal_lookup = {entry["CodeID"]: entry for entry in meals}

    # Step 4: Build the Excel schedule export file.
    excel_data, filename = build_schedule_export_file(
        dates=date_table,
        schedule_entries=schedules,
        day_lookup=day_lookup,
        meal_type_lookup=meal_type_lookup,
        meal_lookup=meal_lookup
    )

    # Step 5: Build and return the final error table using the appearance data and rule data.
    return build_error_table(code_appearances, item_table,
                             {entry["CodeID"]: entry for entry in rules}), excel_data, filename


def gather_data(date_table):
    # Step 1: Gather the day data and matching schedules for the selected dates and current user.
    day_ids, days = gather_day_data(date_table)
    schedules = return_all_schedules({
        "DayID": {"$in": day_ids},
        "UserID": st.session_state.current_user
    })

    # Step 2: Gather all linked meal, category, ingredient, and rule data needed for error checking.
    meal_ids, meal_types, meals = gather_meal_type_meal_data(schedules)
    categories = gather_category_data(meals)
    ingredient_ids_by_meal_id, ingredients = gather_meal_combination_ingredient_data(meal_ids)
    rules = gather_rule_data(categories, ingredients, meal_types, meals)

    # Step 3: Return all gathered data tables for later processing.
    return categories, days, ingredient_ids_by_meal_id, ingredients, meal_types, meals, rules, schedules


def gather_rule_data(categories, ingredients, meal_types, meals):
    # Step 1: Initialise the set used to collect every RuleID attached to the gathered items.
    rule_ids = set()

    # Step 2: Check each gathered collection for attached RuleIDs and gather the matching rule entries.
    for meal_type in meal_types:
        if meal_type.get("RuleID") is not None:
            rule_ids.add(meal_type["RuleID"])
    for meal in meals:
        if meal.get("RuleID") is not None:
            rule_ids.add(meal["RuleID"])
    for category in categories:
        if category.get("RuleID") is not None:
            rule_ids.add(category["RuleID"])
    for ingredient in ingredients:
        if ingredient.get("RuleID") is not None:
            rule_ids.add(ingredient["RuleID"])

    rules = return_all_rules({"CodeID": {"$in": list(rule_ids)}})

    # Step 3: Return the gathered rule entries.
    return rules


def gather_meal_combination_ingredient_data(meal_ids):
    # Step 1: Gather all meal combination rows linked to the selected meals.
    meal_combinations = return_all_meal_combinations({"MealID": {"$in": list(meal_ids)}})

    # Step 2: Build the ingredient lookup by meal and gather all matching ingredient entries.
    ingredient_ids = {
        combo["IngredientID"]
        for combo in meal_combinations
        if combo.get("IngredientID") is not None
    }

    ingredient_ids_by_meal_id = {}
    for combo in meal_combinations:
        meal_id = combo["MealID"]
        ingredient_id = combo["IngredientID"]
        ingredient_ids_by_meal_id.setdefault(meal_id, []).append(ingredient_id)

    ingredients = return_all_ingredients({"CodeID": {"$in": list(ingredient_ids)}})

    # Step 3: Return the meal-to-ingredient lookup and the gathered ingredient entries.
    return ingredient_ids_by_meal_id, ingredients


def gather_category_data(meals):
    # Step 1: Gather the CategoryIDs attached to the selected meals.
    category_ids = {
        meal["CategoryID"]
        for meal in meals
        if meal.get("CategoryID") is not None
    }

    # Step 2: Gather all matching category entries.
    categories = return_all_categories({"CodeID": {"$in": list(category_ids)}})

    # Step 3: Return the gathered category entries.
    return categories


def gather_meal_type_meal_data(schedules):
    # Step 1: Initialise the sets used to collect MealTypeIDs and MealIDs from schedules.
    meal_type_ids = set()
    meal_ids = set()

    # Step 2: Go through schedules and gather the matching meal type and meal entries.
    for schedule in schedules:
        if schedule.get("MealTypeID"):
            meal_type_ids.add(schedule["MealTypeID"])
        if schedule.get("MealID"):
            meal_ids.add(schedule["MealID"])

    meal_types = return_all_meal_types({"CodeID": {"$in": list(meal_type_ids)}})
    meals = return_all_meals({"CodeID": {"$in": list(meal_ids)}})

    # Step 3: Return the gathered MealIDs, meal types, and meals.
    return meal_ids, meal_types, meals


def gather_day_data(date_table):
    # Step 1: Gather all day entries that match the selected date table.
    days = return_all_days({"Date": {"$in": date_table}})

    # Step 2: Build the list of Day CodeIDs from the gathered day entries.
    day_ids = [day["CodeID"] for day in days]

    # Step 3: Return the DayIDs and gathered day entries.
    return day_ids, days


def get_meal_type_dates(meal_types, schedules, table, date_by_day_id):
    """
    Returns a dictionary:
    {
        meal_type_code_id: [date, date, ...]
    }
    for all meal types that appear in the given schedules.
    """
    # Step 1: Build the set of valid MealType CodeIDs for quick matching.
    valid_ids = {meal_type["CodeID"] for meal_type in meal_types}

    # Step 2: Go through schedules and record each appearance date under the matching meal type.
    for schedule in schedules:
        meal_type_id = schedule.get("MealTypeID")
        date_string = date_by_day_id.get(schedule.get("DayID"))

        if meal_type_id in valid_ids and date_string:
            key = ("MealType", meal_type_id)
            table.setdefault(key, []).append(date_string)

    # Step 3: Return the updated appearance table.
    return table


def get_meal_dates(meals, schedules, table, date_by_day_id):
    """
    Returns a dictionary:
    {
        meal_code_id: [date, date, ...]
    }
    for all meals that appear in the given schedules.
    """
    # Step 1: Build the set of valid Meal CodeIDs for quick matching.
    valid_ids = {meal["CodeID"] for meal in meals}

    # Step 2: Go through schedules and record each appearance date under the matching meal.
    for schedule in schedules:
        meal_id = schedule.get("MealID")
        date_string = date_by_day_id.get(schedule.get("DayID"))

        if meal_id in valid_ids and date_string:
            key = ("Meal", meal_id)
            table.setdefault(key, []).append(date_string)

    # Step 3: Return the updated appearance table.
    return table


def get_category_dates(categories, meals, schedules, table, date_by_day_id):
    """
    Returns a dictionary:
    {
        category_code_id: [date, date, ...]
    }
    for all categories that appear through scheduled meals.
    """
    # Step 1: Build the valid CategoryID set and the meal lookup table.
    valid_ids = {category["CodeID"] for category in categories}
    meals_by_id = {meal["CodeID"]: meal for meal in meals}

    # Step 2: Go through schedules, resolve each meal, and record the date under the matching category.
    for schedule in schedules:
        meal = meals_by_id.get(schedule.get("MealID"))
        date_string = date_by_day_id.get(schedule.get("DayID"))

        if not meal or not date_string:
            continue

        category_id = meal.get("CategoryID")

        if category_id in valid_ids:
            key = ("Category", category_id)
            table.setdefault(key, []).append(date_string)

    # Step 3: Return the updated appearance table.
    return table


def get_ingredient_dates(ingredients, ingredient_ids_by_meal_id, schedules, table, date_by_day_id):
    """
    Returns a dictionary:
    {
        ingredient_code_id: [date, date, ...]
    }
    for all ingredients that appear through scheduled meals.
    """
    # Step 1: Build the set of valid Ingredient CodeIDs for quick matching.
    valid_ids = {ingredient["CodeID"] for ingredient in ingredients}

    # Step 2: Go through schedules, resolve each meal's ingredients, and record the date for each ingredient.
    for schedule in schedules:
        meal_id = schedule.get("MealID")
        date_string = date_by_day_id.get(schedule.get("DayID"))

        if not meal_id or not date_string:
            continue

        ingredient_ids = ingredient_ids_by_meal_id.get(meal_id, [])

        for ingredient_id in ingredient_ids:
            if ingredient_id in valid_ids:
                key = ("Ingredient", ingredient_id)
                table.setdefault(key, []).append(date_string)

    # Step 3: Return the updated appearance table.
    return table


def build_error_table(code_appearances, item_table, rules):
    """
    Builds an error table from appearance dates and assigned rules.
    """
    # Step 1: Initialise the error table and start going through each tracked item appearance.
    error_table = []

    # Step 2: Resolve each item, check its rule, and add any violations to the error table.
    for key, date_list in code_appearances.items():
        source_type, code_id = key

        source_table = item_table.get(source_type)
        if not source_table:
            continue

        entry = source_table.get(code_id)
        if not entry:
            continue

        rule_id = entry.get("RuleID")
        if not rule_id:
            continue

        rule = rules.get(rule_id)
        if not rule:
            continue

        quantity = rule["Quantity"]
        per = rule["Per"]
        name = entry.get("Name", code_id)

        violations = find_date_window_violations(date_list, quantity, per)

        for violation in violations:
            error_table.append(
                f"{source_type} '{name}' appears {violation['Count']} times between "
                f"{violation['StartDate']} and {violation['EndDate']}. "
                f"The limit is {quantity} per {per}."
            )

    # Step 3: Return the completed error table.
    return error_table


def get_rule_days(per):
    """
    Returns the number of days in the rolling rule window.
    """
    # Step 1: Check the rule period that was provided.
    if per == "Day":
        # Step 2: Return the correct window size in days for a Day rule.
        return 1
    if per == "Week":
        # Step 2: Return the correct window size in days for a Week rule.
        return 7
    if per == "Month":
        # Step 2: Return the correct window size in days for a Month rule.
        return 30
    if per == "Year":
        # Step 2: Return the correct window size in days for a Year rule.
        return 365

    # Step 3: Return 0 if the rule period is not recognised.
    return 0


def find_date_window_violations(date_list, quantity, per):
    """
    Checks a sorted list of date strings against a rolling rule window.

    Returns a list of violations:
    [
        {
            "StartDate": ...,
            "EndDate": ...,
            "Count": ...
        }
    ]
    """
    # Step 1: Validate the input list and determine the window size for the rule period.
    violations = []

    if not date_list:
        return violations

    window_days = get_rule_days(per)
    if window_days == 0:
        return violations

    sorted_dates = sorted(datetime.strptime(date_string, "%Y-%m-%d") for date_string in date_list)

    # Step 2: Check each date as the start of a rolling window and record any rule violations.
    for i in range(len(sorted_dates)):
        start_date = sorted_dates[i]
        end_date = start_date + timedelta(days=window_days - 1)

        count = 0
        for j in range(i, len(sorted_dates)):
            if sorted_dates[j] <= end_date:
                count += 1
            else:
                break

        if count > quantity:
            violations.append({
                "StartDate": start_date.strftime("%Y-%m-%d"),
                "EndDate": end_date.strftime("%Y-%m-%d"),
                "Count": count
            })

    # Step 3: Return the list of detected violations.
    return violations


def select_mealID(entry, pointer: int):
    # Step 1: Display the Meal Type label (used as section context / caption)
    st.write(meal_name_label)

    # Step 2: Fetch Meal Type lookup table and plain options list
    #         - frequency_table: label -> CodeID
    #         - frequency_list:  [None, "X per Y", ...]
    meal_table, meal_list = convert_ID_to_content()

    # Step 3: Render selectbox for Meal Type selection
    #         - Pre-selects the current Meal Type if entry exists
    #         - Falls back to None if no Meal Type is attached
    meal = st.selectbox(
        meal_name_label,
        meal_list,
        index=meal_id_to_index(
            entry.get("MealID") if entry else None,
            meal_table,
            meal_list
        ),
        label_visibility="collapsed",
        key=f"mealID_button_{pointer}",
    )

    # Step 4: Translate selected display label back into the stored CodeID
    meal_id = meal_table.get(meal)

    # Step 5: Return the resolved Meal Type CodeID (or None)
    return meal_id


def convert_ID_to_content():
    # Step 1: Fetch all meal entries from the database
    table = return_all_meals({'UserID': st.session_state.current_user})

    # Step 2: Initialize lookup table and selectbox options
    #         - lookup maps display label -> CodeID
    #         - options is a plain list for UI components
    lookup = {}
    options = []

    # Step 3: Build label-to-ID mapping and UI options list
    for entry in table:
        label = entry['Name']
        lookup[label] = entry['CodeID']
        options.append(label)

    # Step 4: Return both structures so UI and logic stay in sync
    return lookup, options


def show_day(entry, meal_types, counter: int, add_slot: bool, alter_slot: bool):
    # Step 1: Render the day heading
    st.subheader(entry['Date'])

    # Step 2: Retrieve all schedule entries for this user and day
    data = return_all_schedules({
        'UserID': st.session_state.current_user,
        'DayID': entry['CodeID']
    })

    # Step 3: Extract the MealTypeIDs that already exist in this day's schedule
    existing_meal_types = return_table("MealTypeID", data, None, False)

    # Step 4: Render each meal type slot for the day
    for meal_type in meal_types:
        counter += 1

        with st.container(border=True):
            # Step 4a: Show an existing slot if this meal type is already scheduled
            if meal_type['CodeID'] in existing_meal_types and alter_slot:
                schedule_entry = next(
                    (schedule for schedule in data if schedule['MealTypeID'] == meal_type['CodeID']),
                    None
                )
                show_existing_slot(entry, schedule_entry, meal_type, counter)

            # Step 4b: Show an empty slot if enabled and no schedule exists for this meal type
            elif add_slot and meal_type['CodeID'] not in existing_meal_types:
                present_empty_slot(entry, meal_type, counter)

            # Step 4c: Otherwise render a blank placeholder for alignment
            else:
                show_blank(meal_type)

    # Step 5: Return updated widget counter
    return counter


def show_existing_slot(entry, data, meal_type, counter: int):
    # Step 1: Render the meal type heading
    st.subheader(meal_type['Name'])

    # Step 2: Validate the currently assigned meal
    message, meal_entry, status = validate_meal(data['MealID'])

    if not status:
        # Step 2a: Show validation message if meal lookup fails
        st.write(message)

    elif data['Outcome'] in outcome_table_with_tags["Upcoming"]:
        # Step 2b: Allow the user to update the slot if it is not locked
        mealID = select_mealID(data, counter)
        st.button(
            "Update Slot",
            use_container_width=True,
            on_click=alter_schedule_officially,
            args=[
                data['CodeID'],
                data['UserID'],
                data['MealTypeID'],
                mealID,
                data['DayID'],
                data['Outcome'],
                data['Notes']
            ],
            key=f"update_{entry['CodeID']}_{meal_type['CodeID']}"
        )
        st.button(
            "Remove Slot",
            use_container_width=True,
            on_click=remove_schedule_officially,
            args=[data['CodeID'], data['UserID']],
            key=f"remove_{entry['CodeID']}_{meal_type['CodeID']}"
        )


    else:
        # Step 2c: Show locked meal name and allow the slot to be removed
        st.write(meal_entry[0]['Name'])
        st.button(
            "Remove Slot",
            use_container_width=True,
            on_click=remove_schedule_officially,
            args=[data['CodeID'], data['UserID']],
            key=f"remove_{entry['CodeID']}_{meal_type['CodeID']}"
        )


def present_empty_slot(entry, meal_type, counter: int):
    # Step 1: Render the meal type heading
    st.subheader(meal_type['Name'])

    # Step 2: Let the user choose a meal for this empty slot
    mealID = select_mealID(None, counter)

    # Step 3: Render button to create the new schedule slot
    st.button(
        "Add Slot",
        use_container_width=True,
        on_click=add_schedule_officially,
        args=[
            st.session_state.current_user,
            meal_type['CodeID'],
            mealID,
            entry['CodeID'],
            outcome_table_with_tags['Upcoming'][0],
            ""
        ],
        key=f"add_{entry['CodeID']}_{meal_type['CodeID']}"
    )


def show_blank(meal_type):
    # Step 1: Render the meal type heading
    st.subheader(meal_type['Name'])

    # Step 2: Render empty space to preserve layout consistency
    st.write("")
    st.write("")
    st.write("")


def today_schedule():
    st.header("Meal Schedule")
    add_slot_column, alter_slot_column, range_column = st.columns(3, vertical_alignment="center")
    with add_slot_column:
        add_slot = st.checkbox("Show Empty Slots")
    with alter_slot_column:
        alter_slot = st.checkbox("Show Existing Slots")
    with range_column:
        range_days = st.number_input("Enter Range Value", min_value=0, value=3)
    meal_types = return_all_meal_types({'UserID': st.session_state.current_user})
    meal_types = sorted(meal_types, key=lambda x: x['Priority'])
    table = []
    counter = 0
    pointer = 0
    date_table = generate_date_range(range_days)
    match_date_with_ID = generate_table(date_table)
    for entry in match_date_with_ID:
        counter += 1
        table.append(entry)

        if counter == 3:
            cols = st.columns(len(table))
            for i, val in enumerate(table):
                with cols[i]:
                    pointer = show_day(table[i], meal_types, pointer, add_slot, alter_slot)
            table = []
            counter = 0
    # Step 7a: Render any remaining days that did not fill a row with four
    if len(table) >= 1:
        cols = st.columns(len(table))
        for i, val in enumerate(table):
            with cols[i]:
                pointer = show_day(table[i], meal_types, pointer, add_slot, alter_slot)


def generate_date_range(days):
    # Step 1: Get today's date
    today = datetime.today()

    # Step 2: Build a list of consecutive date strings
    date_table = [
        (today + timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range(days)
    ]

    # Step 3: Return the generated date table
    return date_table


def generate_table(date_table):
    # Step 1: Prepare the output table and gather all existing Day entries for the selected dates.
    match_date_with_id = []

    existing_days = return_all_days({"Date": {"$in": date_table}})
    existing_days_by_date = {entry["Date"]: entry for entry in existing_days}

    # Step 2: Go through each selected date and either reuse or create the matching Day entry.
    for entry in date_table:
        date_entry = existing_days_by_date.get(entry)

        if date_entry:
            # Step 2a: Reuse the existing Day entry.
            match_date_with_id.append({
                "Date": entry,
                "CodeID": date_entry["CodeID"]
            })
        else:
            # Step 2b: Create a new Day entry when no match exists yet.
            message, new_entry, status = create_day(entry, st.session_state.current_user)

            if not status:
                # Step 2c: Store the creation error for global display.
                st.session_state.error = message
                st.session_state.error_status = status
            else:
                # Step 2d: Add the newly created Day entry to the output table.
                match_date_with_id.append({
                    "Date": entry,
                    "CodeID": new_entry["CodeID"]
                })

    # Step 3: Return the completed date-to-CodeID table.
    return match_date_with_id


def update_statuses():
    # Step 1: Get today's date and prepare tracking variables
    today = date.today()
    counter = 0
    table = []

    # Step 2: Check that both the source and target status groups exist
    if outcome_table_with_tags.get('Upcoming') and outcome_table_with_tags.get('Current'):

        # Step 2a: Fetch this user's schedules that currently fall under the Upcoming tag
        schedules = return_all_schedules({
            'Outcome': {'$in': outcome_table_with_tags['Upcoming']},
            'UserID': st.session_state.current_user
        })

        # Step 3: Process each matching schedule
        for schedule in schedules:
            day_date = None

            # Step 3a: Check whether this schedule's DayID has already been looked up
            for row in table:
                if row['CodeID'] == schedule['DayID']:
                    day_date = row['Date']

            # Step 3b: If not already cached, fetch the Day entry and store its date
            if day_date is None:
                day_data = return_all_days({'CodeID': schedule['DayID']})
                if len(day_data) == 1:
                    table.append({
                        'CodeID': schedule['DayID'],
                        'Date': day_data[0]['Date']
                    })
                    day_date = day_data[0]['Date']

            # Step 3c: If a valid day date was found, compare it to today
            if day_date is not None:
                day_date = datetime.strptime(day_date, "%Y-%m-%d").date()

                # Step 3d: If the schedule date is before today, move it to the Current status
                if day_date < today:
                    message, entry, status = update_schedule(
                        schedule['CodeID'],
                        st.session_state.current_user,
                        schedule['MealTypeID'],
                        schedule['MealID'],
                        schedule['DayID'],
                        outcome_table_with_tags['Current'][0],
                        schedule['Notes']
                    )

                    # Step 3e: Count successful updates, stop early if an update fails
                    if status:
                        counter += 1
                    else:
                        st.session_state.error, st.session_state.error_status = message, status
                        return counter

    # Step 4: Return the number of successfully updated schedules
    return counter


def make_schedule_export_filename(dates):
    """
    dates: list of date strings or date objects
    """

    # Step 1: Handle empty input
    if not dates:
        return "schedule_export.xlsx"

    # Step 2: Normalize dates to strings
    dates = [str(d) for d in dates]

    # Step 3: Build filename
    if len(dates) == 1:
        return f"schedule_{dates[0]}.xlsx"

    return f"schedule_{dates[0]}_to_{dates[-1]}.xlsx"


def group_schedule_data(schedule_entries, day_lookup, meal_type_lookup, meal_lookup, dates):
    """
    Returns:
        sorted_meal_types: list of meal type dicts sorted by priority
        planned_data: nested dict {meal_type_name: {date_str: "Meal A\\nMeal B"}}
    """

    # Step 1: Normalize date values once
    date_strings = [str(d) for d in dates]

    # Step 2: Collect every meal type used in the selected schedules
    used_meal_type_ids = set()
    for entry in schedule_entries:
        day_data = day_lookup.get(entry.get("DayID"))
        meal_type_id = entry.get("MealTypeID")

        if not day_data or meal_type_id not in meal_type_lookup:
            continue

        date_value = str(day_data.get("Date"))
        if date_value in date_strings:
            used_meal_type_ids.add(meal_type_id)

    # Step 3: Sort meal types by priority
    sorted_meal_types = sorted(
        [meal_type_lookup[m_id] for m_id in used_meal_type_ids],
        key=lambda x: (x.get("Priority", 999999), x.get("Name", ""))
    )

    # Step 4: Build grouped meal names by meal type and date
    grouped = defaultdict(lambda: defaultdict(list))

    for entry in schedule_entries:
        day_data = day_lookup.get(entry.get("DayID"))
        meal_type_data = meal_type_lookup.get(entry.get("MealTypeID"))
        meal_data = meal_lookup.get(entry.get("MealID"))

        if not day_data or not meal_type_data or not meal_data:
            continue

        date_value = str(day_data.get("Date"))
        if date_value not in date_strings:
            continue

        meal_type_name = meal_type_data.get("Name", "Unknown Meal Type")
        meal_name = meal_data.get("Name", "Unknown Meal")

        grouped[meal_type_name][date_value].append(meal_name)

    # Step 5: Fill final table so every used meal type has every date column
    planned_data = {}
    for meal_type in sorted_meal_types:
        meal_type_name = meal_type.get("Name", "Unknown Meal Type")
        planned_data[meal_type_name] = {}

        for date_value in date_strings:
            meals = grouped[meal_type_name][date_value]
            planned_data[meal_type_name][date_value] = "\n".join(meals) if meals else ""

    return sorted_meal_types, planned_data


def style_schedule_sheet(ws, total_date_columns, data_row_height=None):
    # Step 1: Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Step 2: Style first column and optional row height
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)

        if data_row_height is not None:
            ws.row_dimensions[row].height = data_row_height

    # Step 3: Set widths / wrapping / freeze panes
    ws.column_dimensions["A"].width = 22

    for col in range(2, total_date_columns + 2):
        ws.column_dimensions[get_column_letter(col)].width = 25

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=total_date_columns + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "B2"


def build_planned_sheet(ws, dates, sorted_meal_types, planned_data):
    # Step 1: Add headers
    headers = ["Meal Type"] + [str(d) for d in dates]
    ws.append(headers)

    # Step 2: Add planned schedule rows
    for meal_type in sorted_meal_types:
        meal_type_name = meal_type.get("Name", "Unknown Meal Type")
        row = [meal_type_name]

        for date_value in [str(d) for d in dates]:
            row.append(planned_data[meal_type_name].get(date_value, ""))

        ws.append(row)

    # Step 3: Apply sheet styling
    style_schedule_sheet(ws, len(dates))


def build_actual_sheet(ws, dates, sorted_meal_types):
    # Step 1: Add headers
    headers = ["Meal Type"] + [str(d) for d in dates]
    ws.append(headers)

    # Step 2: Add blank rows for user input
    for meal_type in sorted_meal_types:
        meal_type_name = meal_type.get("Name", "Unknown Meal Type")
        row = [meal_type_name] + ["" for _ in dates]
        ws.append(row)

    # Step 3: Apply sheet styling with taller rows for writing
    table = ["Weight", "Sleep", "Activity", "Alcohol"]
    for entry in table:
        ws.append([entry])
    style_schedule_sheet(ws, len(dates), data_row_height=60)


def build_schedule_workbook(dates, schedule_entries, day_lookup, meal_type_lookup, meal_lookup):
    """
    day_lookup example:
    {
        "day_1": {"CodeID": "day_1", "Date": "2026-04-05"},
        ...
    }

    meal_type_lookup example:
    {
        "mt_1": {"CodeID": "mt_1", "Name": "Breakfast", "Priority": 1},
        "mt_2": {"CodeID": "mt_2", "Name": "Lunch", "Priority": 2},
    }

    meal_lookup example:
    {
        "meal_1": {"CodeID": "meal_1", "Name": "Eggs"},
        "meal_2": {"CodeID": "meal_2", "Name": "Soup"},
    }
    """

    # Step 1: Build grouped planned data
    sorted_meal_types, planned_data = group_schedule_data(
        schedule_entries=schedule_entries,
        day_lookup=day_lookup,
        meal_type_lookup=meal_type_lookup,
        meal_lookup=meal_lookup,
        dates=dates
    )

    # Step 2: Create workbook and sheets
    wb = Workbook()

    ws_planned = wb.active
    ws_planned.title = "Planned Schedule"
    build_planned_sheet(ws_planned, dates, sorted_meal_types, planned_data)

    ws_actual = wb.create_sheet("Actual Intake")
    build_actual_sheet(ws_actual, dates, sorted_meal_types)

    # Step 3: Return workbook
    return wb


def build_schedule_export_file(dates, schedule_entries, day_lookup, meal_type_lookup, meal_lookup):
    # Step 1: Build workbook
    wb = build_schedule_workbook(
        dates=dates,
        schedule_entries=schedule_entries,
        day_lookup=day_lookup,
        meal_type_lookup=meal_type_lookup,
        meal_lookup=meal_lookup
    )

    # Step 2: Build filename
    filename = make_schedule_export_filename(dates)

    # Step 3: Save to memory for download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, filename
