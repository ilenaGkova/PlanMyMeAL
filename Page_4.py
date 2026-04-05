import streamlit as st
from Day import return_all_days
from Ingredient import validate_ingredient
from Meal import validate_meal
from MealCombination import return_all_meal_combinations
from Menu import menu
from Page_6 import get_table
from Schedule import return_all_schedules
from UnitType import validate_unit_type
from User import validate_user
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


# Step 1: Initialize session state variables (first run only)

if 'page' not in st.session_state:
    st.session_state.page = 1  # Default page to load

if 'error_status' not in st.session_state:
    st.session_state.error_status = None  # Tracks whether an error message should be shown

if 'current_user' not in st.session_state:
    st.session_state.current_user = None  # Stores the CodeID of the signed-in user

if 'error' not in st.session_state:
    st.session_state.error = 'You are doing great! Keep going.'  # Stores the current error message


def page_4_layout():
    # Step 1: Validate current user
    # Ensures the user exists and is allowed to access this page
    message, entry, status = validate_user(st.session_state.current_user)

    if status:
        # Step 2: Render page UI and process grocery data

        # Step 2a: Render main navigation/menu
        menu(entry)

        # Step 2b: Page title
        st.title("My grocery List")

        # Step 3: Date selection input
        st.header("Date Selection")
        date_table = get_table()

        if len(date_table) >= 1:

            # Step 4: Build meal + ingredient data and display meal breakdown

            # Step 4a: Convert selected dates → aggregated meals
            meal_table = make_meal_table(date_table)

            if len(meal_table) >= 1:

                with st.container(border=True):

                    st.header("This grocery list corresponds to:")

                    # Step 4b: Convert meals → ingredients (also displays meal summary)
                    ingredient_table = make_ingredient_table(meal_table)

                    # Step 5: Render final grocery list
                    with st.container(border=True):
                        st.header("Grocery List")
                        for entry in ingredient_table:
                            st.write(f"🛒 {entry['Quantity']} {entry['Unit']}(s) of {entry['Name']}")

                # Step 6: Make exel file of ingredient table
                excel_data, filename = build_ingredient_export_file(ingredient_table, date_table)

                # Step 6a: Download exel file at will
                st.download_button(
                    label="Download ingredient list",
                    data=excel_data,
                    use_container_width=True,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:

                st.write('You have no meals planed for this duration of time')

    else:
        # Step 7: Handle validation failure
        # Store error for global error display system
        st.session_state.error, st.session_state.error_status = message, status


def make_meal_table(date_table):
    # Step 1: Transform selected dates → meal table

    meal_table = []

    for entry in date_table:
        # Step 2: Resolve date into system Day record
        date_data = return_all_days({'Date': entry})

        # Step 3: If valid day found, fetch scheduled meals
        if len(date_data) == 1:
            data = return_all_schedules({'DayID': date_data[0]['CodeID'], 'UserID': st.session_state.current_user})

            # Step 4: Aggregate meals into meal_table
            for meal in data:
                meal_table = add_to_table(meal_table, meal['MealID'], 1, 'Meal')

    # Step 1d: Return aggregated meals
    return meal_table


def make_ingredient_table(meal_table):
    # Step 1: Transform meal table → ingredient table

    ingredient_table = []

    for entry in meal_table:
        # Step 2: Fetch ingredient list for each meal
        data = return_all_meal_combinations({'MealID': entry['CodeID']})

        # Step 3: Display meal breakdown (for user context)
        st.write(f"-> {entry['Quantity']} Portion(s) of {entry['Name']} - {len(data)} Ingredients Each")

        # Step 4: Aggregate ingredients (scaled by meal quantity)
        for ingredient in data:
            ingredient_table = add_to_table(
                ingredient_table,
                ingredient['IngredientID'],
                ingredient['Quantity'] * entry['Quantity'],
                'Ingredient'
            )

    # Step 5: Return aggregated ingredients
    return ingredient_table


# Mapping item types to their validation functions
validate_item = {
    'Meal': validate_meal,
    'Ingredient': validate_ingredient
}


def add_to_table(table, entry: str, quantity: float, item: str):
    # Step 1: Check if entry already exists in table
    for row in table:
        if row['CodeID'] == entry:
            row['Quantity'] += quantity
            return table

    # Step 2: Create new entry if not found
    new_row = {'CodeID': entry, 'Quantity': quantity, 'Name': None, 'Unit': None}

    # Step 3: Retrieve name using validation function
    message, data, status = validate_item[item](entry)
    if not status:
        new_row['Name'] = message
    else:
        new_row['Name'] = data[0]['Name']

    # Step 4: Retrieve unit only for ingredients
    if item == 'Ingredient':
        message, unitType, status = validate_unit_type(data[0]['UnitTypeID'])
        if not status:
            new_row['Unit'] = message
        else:
            new_row['Unit'] = unitType[0]['Name']

    # Step 5: Add new entry to table
    table.append(new_row)
    return table


def make_export_filename(dates):
    """
    dates: list of date strings or date objects
    """

    # Step 1: Handle empty input (fallback filename)
    if not dates:
        return "ingredients_export.xlsx"

    # Step 2: Normalize dates to strings
    dates = [str(d) for d in dates]

    # Step 3: Build filename based on number of dates
    if len(dates) == 1:
        return f"ingredients_{dates[0]}.xlsx"

    return f"ingredients_{dates[0]}_to_{dates[-1]}.xlsx"


def build_ingredient_workbook(ingredient_rows):
    """
    ingredient_rows: list of dicts like
    {
        'CodeID': entry,
        'Quantity': quantity,
        'Name': ingredient_name,
        'Unit': unit
    }
    """

    # Step 1: Create workbook and set up sheet + headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingredients"

    headers = ["Ingredient", "Quantity", "Unit", "Existing", "Outcome"]
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Step 2: Fill rows with data + formulas
    row_num = 2
    for item in ingredient_rows:
        ws.cell(row=row_num, column=1, value=item.get("Name"))
        ws.cell(row=row_num, column=2, value=item.get("Quantity"))
        ws.cell(row=row_num, column=3, value=item.get("Unit"))
        ws.cell(row=row_num, column=4, value=None)  # Existing stays blank
        ws.cell(row=row_num, column=5, value=f"=D{row_num}-B{row_num}")  # Outcome
        row_num += 1

    # Step 3: Format sheet
    widths = {
        1: 30,
        2: 12,
        3: 12,
        4: 12,
        5: 12,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Optional alignment for data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Conditional formatting for full row
    if ingredient_rows:
        last_row = len(ingredient_rows) + 1
        full_range = f"A2:E{last_row}"

        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        yellow_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

        ws.conditional_formatting.add(
            full_range,
            FormulaRule(
                formula=["$E2<0"],
                fill=red_fill
            )
        )

        ws.conditional_formatting.add(
            full_range,
            FormulaRule(
                formula=["$E2=0"],
                fill=yellow_fill
            )
        )

    return wb


def build_ingredient_export_file(ingredient_rows, dates):
    # Step 1: Build the Excel workbook
    wb = build_ingredient_workbook(ingredient_rows)

    # Step 2: Generate the export filename
    filename = make_export_filename(dates)

    # Step 3: Save workbook to in-memory file (for download)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, filename